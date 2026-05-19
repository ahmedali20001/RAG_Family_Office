"""
SEC IAPD Family Office Data Extractor
======================================
Methodology:
  Discovery : CRD numbers pre-identified in crd.csv
  Extraction: SEC IAPD public REST API  https://api.adviserinfo.sec.gov/search/individual/{CRD}
  Enrichment: Years of experience (days-in-industry), license counts, exam categories,
               current/previous firm lists, state registrations, disclosure counts
  Validation : Field-presence checks, type checks, CRD numeric check, scope checks,
               confidence scoring per record
  Output    : family_offices.xlsx  (3 sheets: Records, Validation_Chain, Methodology)
"""

import csv, json, ssl, time, datetime, re, urllib.request
from typing import Optional, List
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── config ──────────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent
CRD_CSV    = BASE_DIR / "crd.csv"
OUT_XLSX   = BASE_DIR / "family_offices_v2.xlsx"
API_BASE   = "https://api.adviserinfo.sec.gov/search/individual/{crd}"
PDF_BASE   = "https://reports.adviserinfo.sec.gov/reports/individual/individual_{crd}.pdf"
WEB_BASE   = "https://adviserinfo.sec.gov/individual/summary/{crd}"
DELAY_SEC  = 1.2          # polite crawl delay
VALIDATION_CHAIN_CRDS = []  # will be filled with first 3 successful CRDs

# ── SSL context (SEC cert sometimes trips on older Python bundles) ────────────
ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode    = ssl.CERT_NONE

HEADERS = {"User-Agent": "FamilyOfficeResearch/1.0 (assessment@example.com)"}

# ── helpers ──────────────────────────────────────────────────────────────────

def fetch_individual(crd: str) -> Optional[dict]:
    """Fetch raw JSON from IAPD API. Returns parsed dict or None on failure and saves raw json."""
    url = API_BASE.format(crd=crd)
    req = urllib.request.Request(url, headers=HEADERS)
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=15) as r:
            raw = json.loads(r.read())
        hits = raw.get("hits", {}).get("hits", [])
        if not hits:
            return None
        
        iacontent_str = hits[0]["_source"]["iacontent"]
        parsed = json.loads(iacontent_str)
        
        # Save the raw JSON data to a folder
        json_dir = BASE_DIR / "raw_json"
        json_dir.mkdir(parents=True, exist_ok=True)
        with open(json_dir / f"{crd}.json", "w", encoding="utf-8") as f:
            json.dump(parsed, f, indent=2, ensure_ascii=False)
            
        return parsed
    except Exception as exc:
        print(f"  [WARN] CRD {crd}: {exc}")
        return None


def years_of_experience(industry_date: str) -> int:
    """Calculate years of experience matching the SEC IAPD Angular UI.
    
    The UI takes the daysInIndustryCalculatedDateIAPD and rounds it.
    For 7/2/2010 (approx 15.88 years), it displays 16.
    Returns 0 if the date is missing or unparseable.
    """
    if not industry_date:
        return 0
    try:
        dt    = datetime.datetime.strptime(industry_date.strip(), "%m/%d/%Y")
        delta = datetime.datetime.today() - dt
        # round the floating point years just like the SEC JS frontend does
        return round(delta.days / 365.25)
    except Exception:
        return 0


def earliest_employment_date(data: dict) -> str:
    """Fallback: find the earliest registrationBeginDate across all employments."""
    all_emp = (
        data.get("currentEmployments", []) +
        data.get("previousEmployments", []) +
        data.get("currentIAEmployments", []) +
        data.get("previousIAEmployments", [])
    )
    dates = []
    for e in all_emp:
        d = e.get("registrationBeginDate", "")
        try:
            dates.append(datetime.datetime.strptime(d.strip(), "%m/%d/%Y"))
        except Exception:
            pass
    if not dates:
        return ""
    return min(dates).strftime("%m/%d/%Y")


def license_types(ia_scope: str, bc_scope: str) -> str:
    types = []
    if str(ia_scope).upper() == "ACTIVE":
        types.append("IA")
    if str(bc_scope).upper() == "ACTIVE":
        types.append("B")
    return ", ".join(types) if types else "N/A"


def primary_location(employments: list) -> tuple[str, str, str]:
    """Return (city, state, country) from first employment with a location."""
    for emp in employments:
        for branch in emp.get("branchOfficeLocations", []):
            city    = branch.get("city", "")
            state   = branch.get("state", "")
            country = branch.get("country", "United States")
            if city or state:
                return city, state, country
    return "", "", ""


def format_firms(employments: list) -> str:
    parts = []
    for e in employments:
        name  = e.get("firmName", "")
        start = e.get("registrationBeginDate", "")
        end   = e.get("registrationEndDate", "")
        date  = f"{start}–{end}" if end else f"{start}–present"
        parts.append(f"{name} ({date})")
    return "; ".join(parts)


def format_states(states: list) -> str:
    seen, out = set(), []
    for s in states:
        k = s.get("state", "")
        if k and k not in seen:
            seen.add(k)
            out.append(k)
    return ", ".join(out)


def format_exams(exams: list) -> str:
    """Format exam list into 'Name (Date)' string."""
    if not exams:
        return "None"
    return "; ".join([f"{e.get('examName','Unknown')} ({e.get('examTakenDate','')})" for e in exams])

def format_disclosures(disclosures: list) -> str:
    """Format detailed disclosure data into a readable block dynamically, matching SEC UI."""
    if not disclosures:
        return "None"
    
    formatted = []
    for i, d in enumerate(disclosures, 1):
        detail = d.get("disclosureDetail", {})
        parts = [f"--- Disclosure #{i} ({d.get('disclosureType', 'Unknown')} - {d.get('disclosureResolution', 'Unknown')}) ---"]
        
        # Include the date if available
        event_date = d.get("eventDate")
        if event_date:
            parts.append(f"Date: {event_date}")
            
        if detail:
            # Order the keys logically so they display like the website
            known_order = [
                "Initiated By", 
                "DocketNumberFDA", 
                "DocketNumberAAO", 
                "Allegations", 
                "Resolution", 
                "SanctionDetails", 
                "Regulator Statement", 
                "Broker Comment"
            ]
            # Capture any other keys dynamically that we didn't specify
            all_keys = known_order + [k for k in detail.keys() if k not in known_order]
            
            for k in all_keys:
                if k not in detail:
                    continue
                v = detail[k]
                if not v:
                    continue
                
                # Normalize key names for the output
                display_key = k
                if k == "DocketNumberFDA":
                    display_key = "Docket Number (FDA)"
                elif k == "DocketNumberAAO":
                    display_key = "Docket Number (AAO)"
                
                # Format the values depending on type
                if k == "SanctionDetails":
                    for s in v:
                        if s.get("Sanctions"):
                            parts.append(f"Sanctions: {s.get('Sanctions')}")
                        for sub in s.get("SanctionDetails", []):
                            if sub.get("Amount"):
                                parts.append(f"Amount: {sub.get('Amount')}")
                elif k == "Broker Comment":
                    if isinstance(v, list):
                        parts.append(f"Broker Comment: " + " ".join(str(x) for x in v))
                    else:
                        parts.append(f"Broker Comment: {v}")
                else:
                    if isinstance(v, list):
                        parts.append(f"{display_key}: " + " ".join(str(x) for x in v))
                    else:
                        parts.append(f"{display_key}: {v}")
        else:
            parts.append("No details found in API.")
            
        formatted.append("\n".join(parts))
        
    return "\n\n".join(formatted)


# ── validation logic ─────────────────────────────────────────────────────────

REQUIRED_FIELDS = ["firstName", "lastName", "individualId"]

def validate_record(crd: str, data: Optional[dict]) -> dict:
    """
    Returns a validation dict with:
      - checks: list of (check_name, pass/fail, detail)
      - confidence: 0–100 score
      - issues: list of problems
    """
    checks = []
    issues = []

    # 1. API reachability
    checks.append(("API Reachable", data is not None, "200 OK from IAPD API" if data else "No data returned"))
    if data is None:
        return {"checks": checks, "confidence": 0, "issues": ["Record not found"]}

    bi = data.get("basicInformation", {})

    # 2. CRD numeric format
    crd_ok = str(crd).isdigit()
    checks.append(("CRD Format", crd_ok, f"CRD={crd} is numeric" if crd_ok else "Non-numeric CRD"))
    if not crd_ok:
        issues.append("Invalid CRD format")

    # 3. CRD matches returned record
    returned_id = str(bi.get("individualId", ""))
    id_match    = returned_id == str(crd)
    checks.append(("CRD Match", id_match, f"API returned id={returned_id}"))
    if not id_match:
        issues.append(f"CRD mismatch: requested {crd}, got {returned_id}")

    # 4. Required fields present
    for field in REQUIRED_FIELDS:
        present = bool(bi.get(field))
        checks.append((f"Field:{field}", present, bi.get(field, "MISSING")))
        if not present:
            issues.append(f"Missing required field: {field}")

    # 5. Scope sanity (at least one active scope)
    ia = str(bi.get("iaScope", "")).upper()
    bc = str(bi.get("bcScope", "")).upper()
    scope_ok = ia == "ACTIVE" or bc == "ACTIVE"
    checks.append(("Active Scope", scope_ok, f"iaScope={ia} bcScope={bc}"))
    if not scope_ok:
        issues.append("No active IA or BC scope")

    # 6. Currently Active Employment (API Registration History has a Firm that is "- Present")
    curr_bd = data.get("currentEmployments", [])
    curr_ia = data.get("currentIAEmployments", [])
    is_active = len(curr_bd) > 0 or len(curr_ia) > 0
    checks.append(("Currently Active", is_active, "Has active BD/IA employment (- Present)"))
    if not is_active:
        issues.append("No currently active employment found")

    # 7. Disclosure check (flag vs list consistency)
    disc_flag = data.get("disclosureFlag", "N")
    disc_list = data.get("disclosures", [])
    disc_ok   = (disc_flag == "Y") == (len(disc_list) > 0)
    checks.append(("Disclosure Consistency", disc_ok,
                   f"flag={disc_flag}, list_count={len(disc_list)}"))
    if not disc_ok:
        issues.append("Disclosure flag/list mismatch")

    # 8. Check Licenses
    reg_states = data.get("registeredStates", [])
    has_licenses = len(reg_states) > 0
    checks.append(("Has Licenses", has_licenses, f"{len(reg_states)} registered states"))
    if not has_licenses:
        issues.append("No registered states (Licenses) found")

    # 9. Check Register Locations
    city, state, country = primary_location(curr_bd + curr_ia)
    has_location = bool(city and state)
    checks.append(("Has Location", has_location, f"{city}, {state}"))
    if not has_location:
        issues.append("No primary registration location found")

    # 10. Check Firms
    prev_bd = data.get("previousEmployments", [])
    prev_ia = data.get("previousIAEmployments", [])
    all_employments = curr_bd + curr_ia + prev_bd + prev_ia
    unique_firm_ids = {f.get("firmId") for f in all_employments if f.get("firmId")}
    has_firms = len(unique_firm_ids) > 0
    checks.append(("Has Firms", has_firms, f"{len(unique_firm_ids)} unique firms"))
    if not has_firms:
        issues.append("No firms found in registration history")

    # 11. Check Years of Experience
    industry_date = bi.get("daysInIndustryCalculatedDateIAPD", "").strip()
    if not industry_date:
        industry_date = earliest_employment_date(data)
    yoe = years_of_experience(industry_date)
    has_yoe = yoe > 0
    checks.append(("Has Experience", has_yoe, f"{yoe} Years of Experience"))
    if not has_yoe:
        issues.append("Years of experience is 0 or unparseable")

    # confidence score: each check worth equal weight
    total   = len(checks)
    passed  = sum(1 for _, ok, _ in checks if ok)
    confidence = round((passed / total) * 100)

    return {"checks": checks, "confidence": confidence, "issues": issues}


# ── extraction ───────────────────────────────────────────────────────────────

def extract_record(crd: str, data: dict, validation: dict) -> dict:
    bi   = data.get("basicInformation", {})
    curr = data.get("currentEmployments", [])
    prev = data.get("previousEmployments", [])
    curr_ia = data.get("currentIAEmployments", [])
    prev_ia = data.get("previousIAEmployments", [])
    all_curr = curr_ia if curr_ia else curr
    disc = data.get("disclosures", [])
    exams_count = data.get("examsCount", {})
    state_exams = data.get("stateExamCategory", [])
    prin_exams  = data.get("principalExamCategory", [])
    prod_exams  = data.get("productExamCategory", [])
    reg_states  = data.get("registeredStates", [])
    reg_sros    = data.get("registeredSROs", [])

    city, state, country = primary_location(all_curr)

    # Years of Experience: use SEC's synthetic date; fall back to earliest employment
    industry_date = bi.get("daysInIndustryCalculatedDateIAPD", "").strip()
    if not industry_date:
        industry_date = earliest_employment_date(data)
    yoe = years_of_experience(industry_date)

    full_name   = f"{bi.get('firstName','')} {bi.get('middleName','')} {bi.get('lastName','')}".strip()
    other_names = "; ".join(bi.get("otherNames", []))

    current_firms  = format_firms(curr)
    current_ia_firms = format_firms(curr_ia)
    previous_firms = format_firms(prev)
    previous_ia_firms = format_firms(prev_ia)

    all_sros = "; ".join(s.get("sro", "") for s in reg_sros if s.get("sro"))

    # Firm calculation: unique firmIds across all employments
    all_employments = curr + prev + curr_ia + prev_ia
    unique_firm_ids = {f.get("firmId") for f in all_employments if f.get("firmId")}
    total_firms_count = len(unique_firm_ids)

    # Registration counts
    reg_count = data.get("registrationCount", {})
    total_firms = (
        len(curr) + len(prev)
        + len([f for f in curr_ia if f not in curr])
        + len([f for f in prev_ia if f not in prev])
    )

    return {
        "CRD#"                        : crd,
        "Full Name"                   : full_name,
        "Other Names"                 : other_names,
        "IA Scope"                    : bi.get("iaScope", ""),
        "BC Scope"                    : bi.get("bcScope", ""),
        "License Types"               : license_types(bi.get("iaScope",""), bi.get("bcScope","")),
        "Years of Experience"         : yoe,
        "Total Disclosures"           : len(disc),
        "Disclosure Details"          : format_disclosures(disc),
        "Disclosure Flag"             : data.get("disclosureFlag","N"),
        "IA Disclosure Flag"          : data.get("iaDisclosureFlag","N"),
        "Current Firms (BD)"          : current_firms,
        "Current Firms (IA)"          : current_ia_firms,
        "Total Current BD Firms"      : len(curr),
        "Total Current IA Firms"      : len(curr_ia),
        "Previous Firms (BD)"         : previous_firms,
        "Previous Firms (IA)"         : previous_ia_firms,
        "Total Previous BD Firms"     : len(prev),
        "Total Previous IA Firms"     : len(prev_ia),
        "Total Firms (All)"           : total_firms_count,
        "Primary City"                : city,
        "Primary State"               : state,
        "Primary Country"             : country,
        "Registered States"           : format_states(reg_states),
        "State Registration Count"    : reg_count.get("approvedStateRegistrationCount", 0),
        "IA State Registration Count" : reg_count.get("approvedIAStateRegistrationCount", 0),
        "FINRA Registration Count"    : reg_count.get("approvedFinraRegistrationCount", 0),
        "SRO Registration Count"      : reg_count.get("approvedSRORegistrationCount", 0),
        "Registered SROs"             : all_sros,
        "State Exam Count"            : exams_count.get("stateExamCount", 0),
        "Principal Exam Count"        : exams_count.get("principalExamCount", 0),
        "Product Exam Count"          : exams_count.get("productExamCount", 0),
        "Total Licenses"              : len(reg_states),
        "State Exams"                 : format_exams(state_exams),
        "Principal Exams"             : format_exams(prin_exams),
        "Product Exams"               : format_exams(prod_exams),
        "Has Inactive Registration"   : reg_count.get("hasInactiveRegistration","N"),
        "Has Suspended Registration"  : reg_count.get("hasSuspendedRegistration","N"),
        "SEC Adviser Info URL"        : WEB_BASE.format(crd=crd),
        "PDF Report URL"              : PDF_BASE.format(crd=crd),
        "Confidence Score (%)"        : validation["confidence"],
        "Validation Issues"           : "; ".join(validation["issues"]) if validation["issues"] else "None",
        "Extraction Timestamp"        : datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ── Excel styling helpers ────────────────────────────────────────────────────

def style_header_row(ws, row: int, bg: str = "1F3864"):
    fill   = PatternFill("solid", fgColor=bg)
    font   = Font(color="FFFFFF", bold=True, size=10)
    border = Border(
        bottom=Side(style="medium", color="FFFFFF"),
        right =Side(style="thin",   color="FFFFFF"),
    )
    for cell in ws[row]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border


def auto_width(ws, max_w: int = 50):
    for col_cells in ws.columns:
        length = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, max_w)


def confidence_color(score: int) -> str:
    if score >= 85:
        return "C6EFCE"   # green
    if score >= 60:
        return "FFEB9C"   # yellow
    return "FFC7CE"       # red


# ── Excel writing ─────────────────────────────────────────────────────────────

def write_records_sheet(wb, records: list):
    ws = wb.create_sheet("Records")
    if not records:
        return

    headers = list(records[0].keys())
    ws.append(headers)
    style_header_row(ws, 1)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    alt1 = PatternFill("solid", fgColor="EEF2F7")
    alt2 = PatternFill("solid", fgColor="FFFFFF")

    for i, rec in enumerate(records, start=2):
        row_data = [rec.get(h, "") for h in headers]
        ws.append(row_data)
        fill = alt1 if i % 2 == 0 else alt2
        conf_col = headers.index("Confidence Score (%)") + 1

        for j, cell in enumerate(ws[i], start=1):
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if j == conf_col:
                score = rec.get("Confidence Score (%)", 0)
                cell.fill = PatternFill("solid", fgColor=confidence_color(score))
                cell.font = Font(bold=True)
            else:
                cell.fill = fill

    auto_width(ws)
    
    # Adjust width for Disclosures in the Records sheet to keep it neat
    if "Disclosure Details" in headers:
        col_idx = headers.index("Disclosure Details") + 1
        ws.column_dimensions[get_column_letter(col_idx)].width = 75

    ws.auto_filter.ref = ws.dimensions


def write_validation_sheet(wb, chain_data: list):
    ws = wb.create_sheet("Validation_Chain")
    ws.append(["Validation Chain – 3 Sample Records"])
    ws["A1"].font = Font(bold=True, size=13, color="1F3864")
    ws.merge_cells("A1:G1")
    ws.append([])

    for entry in chain_data:
        crd       = entry["crd"]
        name      = entry["name"]
        rec       = entry["record"]
        val       = entry["validation"]

        # Section header
        ws.append([f"CRD# {crd}  –  {name}"])
        row = ws.max_row
        ws[f"A{row}"].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f"A{row}"].fill = PatternFill("solid", fgColor="1F3864")
        ws.merge_cells(f"A{row}:G{row}")

        # Meta
        meta_rows = [
            ["Discovery Source",   "crd.csv (pre-identified CRD list)"],
            ["Extraction Method",  f"SEC IAPD REST API → {API_BASE.format(crd=crd)}"],
            ["Enrichment Steps",   "Years of experience (date math), license type inference, "
                                   "primary location from branch offices, exam aggregation, "
                                   "state registration counts"],
            ["Confidence Score",   f"{val['confidence']}%"],
            ["Validation Issues",  "; ".join(val['issues']) if val['issues'] else "None"],
            ["SEC Web URL",        WEB_BASE.format(crd=crd)],
            ["PDF Report URL",     PDF_BASE.format(crd=crd)],
        ]
        for mr in meta_rows:
            ws.append([""] + mr)
            r = ws.max_row
            ws[f"B{r}"].font = Font(bold=True)
            ws[f"C{r}"].alignment = Alignment(wrap_text=True)

        ws.append([])

        # Checks table
        ws.append(["", "Check Name", "Pass/Fail", "Detail"])
        r = ws.max_row
        for col_letter, title in [("B","Check Name"),("C","Pass/Fail"),("D","Detail")]:
            ws[f"{col_letter}{r}"].font = Font(bold=True, color="FFFFFF")
            ws[f"{col_letter}{r}"].fill = PatternFill("solid", fgColor="2E75B6")

        for check_name, ok, detail in val["checks"]:
            ws.append(["", check_name, "✔ PASS" if ok else "✘ FAIL", str(detail)])
            r   = ws.max_row
            clr = "C6EFCE" if ok else "FFC7CE"
            ws[f"C{r}"].fill = PatternFill("solid", fgColor=clr)
            ws[f"C{r}"].font = Font(bold=True)

        ws.append([])
        ws.append([])

    # column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20


def write_methodology_sheet(wb, total: int, success: int, failed_crds: list):
    ws = wb.create_sheet("Methodology")
    content = [
        ("METHODOLOGY SUMMARY", None),
        ("", None),
        ("1. DISCOVERY", None),
        ("Source", "50 CRD numbers provided in crd.csv, each mapped to a unique adviser profile on SEC IAPD."),
        ("URL pattern", "https://adviserinfo.sec.gov/individual/summary/{CRD}"),
        ("", None),
        ("2. EXTRACTION", None),
        ("API Endpoint", "https://api.adviserinfo.sec.gov/search/individual/{CRD}"),
        ("Protocol", "HTTPS GET with User-Agent header; SSL verification relaxed for SEC cert chain."),
        ("Parsing", "JSON → hits[0]._source.iacontent (nested JSON string) → Python dict."),
        ("Rate limiting", f"{DELAY_SEC}s polite delay between requests."),
        ("", None),
        ("3. ENRICHMENT", None),
        ("Years of Experience", "Computed from basicInformation.daysInIndustryCalculatedDateIAPD to today."),
        ("License Types", "Derived from iaScope / bcScope flags (IA = Investment Adviser, B = Broker)."),
        ("Primary Location", "Extracted from first branch office of currentIAEmployments or currentEmployments."),
        ("Exam Aggregation", "Counted and listed state, principal, and product exam categories."),
        ("Firm History", "Formatted as Name (start–end) for current and previous BD/IA firms."),
        ("PDF & Web URLs", "Constructed deterministically from CRD number."),
        ("", None),
        ("4. VALIDATION CHECKS", None),
        ("API Reachable", "HTTP 200 and non-empty hits array."),
        ("CRD Format", "CRD must be purely numeric."),
        ("CRD Match", "individualId in response must equal requested CRD."),
        ("Required Fields", "firstName, lastName, individualId must be non-empty."),
        ("Active Scope", "At least one of iaScope or bcScope must be ACTIVE."),
        ("Has Current Employment", "currentEmployments list must be non-empty."),
        ("Disclosure Consistency", "disclosureFlag must agree with disclosures list."),
        ("Experience Date", "daysInIndustryCalculatedDateIAPD must parse as MM/DD/YYYY."),
        ("Confidence Score", "% of checks passed. ≥85% = Green, 60–84% = Yellow, <60% = Red."),
        ("", None),
        ("5. RESULTS", None),
        ("Total CRDs attempted", str(total)),
        ("Records extracted", str(success)),
        ("Failed / not found", str(len(failed_crds))),
        ("Failed CRDs", ", ".join(failed_crds) if failed_crds else "None"),
        ("", None),
        ("6. WHAT I WOULD IMPROVE", None),
        ("Firm classification", "Cross-reference firmId against SEC firm registry to confirm Family Office classification."),
        ("AUM enrichment", "Pull Form ADV Part 1A filings to add AUM, client types, fee structures."),
        ("Contact data", "Enrich with LinkedIn/company websites for direct contact intelligence."),
        ("Re-validation cadence", "Re-run extraction monthly; IAPD data updates with regulatory filings."),
        ("NLP enrichment", "Use LLM to extract investment thesis from Form ADV narrative sections."),
        ("Alert system", "Monitor for new disclosures or registration changes via IAPD change feed."),
    ]

    title_font  = Font(bold=True, size=14, color="1F3864")
    sect_font   = Font(bold=True, size=11, color="FFFFFF")
    sect_fill   = PatternFill("solid", fgColor="1F3864")
    key_font    = Font(bold=True)
    val_align   = Alignment(wrap_text=True, vertical="top")

    sections = {"1. DISCOVERY","2. EXTRACTION","3. ENRICHMENT",
                "4. VALIDATION CHECKS","5. RESULTS","6. WHAT I WOULD IMPROVE"}

    for label, value in content:
        if label == "METHODOLOGY SUMMARY":
            ws.append([label])
            ws[f"A{ws.max_row}"].font = title_font
        elif label in sections:
            ws.append([label, value or ""])
            r = ws.max_row
            ws[f"A{r}"].font = sect_font
            ws[f"A{r}"].fill = sect_fill
            ws.merge_cells(f"A{r}:B{r}")
        elif label == "":
            ws.append([""])
        else:
            ws.append([label, value])
            r = ws.max_row
            ws[f"A{r}"].font  = key_font
            ws[f"B{r}"].alignment = val_align

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 80


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    # Read CRDs
    crds = []
    with open(CRD_CSV, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            crd = str(row.get("CRD_Numbers", "")).strip()
            if crd:
                crds.append(crd)

    print(f"Loaded {len(crds)} CRDs from {CRD_CSV}")

    records      = []
    failed_crds  = []
    chain_data   = []   # validation chain for first 3 successes

    for i, crd in enumerate(crds, 1):
        print(f"[{i:02d}/{len(crds)}] CRD {crd} ...", end=" ", flush=True)
        data       = fetch_individual(crd)
        validation = validate_record(crd, data)

        if data is None:
            print(f"FAILED  (confidence={validation['confidence']}%)")
            failed_crds.append(crd)
            # Still write a placeholder row
            records.append({
                "CRD#"                        : crd,
                "Full Name"                   : "NOT FOUND",
                "Other Names"                 : "",
                "IA Scope"                    : "",
                "BC Scope"                    : "",
                "License Types"               : "",
                "Years of Experience"         : "",
                "Total Disclosures"           : "",
                "Disclosure Details"          : "",
                "Disclosure Flag"             : "",
                "IA Disclosure Flag"          : "",
                "Current Firms (BD)"          : "",
                "Current Firms (IA)"          : "",
                "Total Current BD Firms"      : "",
                "Total Current IA Firms"      : "",
                "Previous Firms (BD)"         : "",
                "Previous Firms (IA)"         : "",
                "Total Previous BD Firms"     : "",
                "Total Previous IA Firms"     : "",
                "Total Firms (All)"           : "",
                "Primary City"               : "",
                "Primary State"               : "",
                "Primary Country"             : "",
                "Registered States"           : "",
                "State Registration Count"    : "",
                "IA State Registration Count" : "",
                "FINRA Registration Count"    : "",
                "SRO Registration Count"      : "",
                "Registered SROs"             : "",
                "State Exam Count"            : "",
                "Principal Exam Count"        : "",
                "Product Exam Count"          : "",
                "Total Licenses"              : "",
                "State Exams"                 : "",
                "Principal Exams"             : "",
                "Product Exams"               : "",
                "Has Inactive Registration"   : "",
                "Has Suspended Registration"  : "",
                "SEC Adviser Info URL"        : WEB_BASE.format(crd=crd),
                "PDF Report URL"              : PDF_BASE.format(crd=crd),
                "Confidence Score (%)"        : 0,
                "Validation Issues"           : "; ".join(validation["issues"]),
                "Extraction Timestamp"        : datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
        else:
            rec = extract_record(crd, data, validation)
            records.append(rec)
            print(f"OK  {rec['Full Name']}  (confidence={validation['confidence']}%)")

            if len(chain_data) < 3:
                chain_data.append({
                    "crd"       : crd,
                    "name"      : rec["Full Name"],
                    "record"    : rec,
                    "validation": validation,
                })

        time.sleep(DELAY_SEC)

    # Build Excel
    print(f"\nWriting {OUT_XLSX} ...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)          # remove default sheet

    write_records_sheet(wb, records)
    write_validation_sheet(wb, chain_data)
    write_methodology_sheet(wb, len(crds), len(records) - len(failed_crds), failed_crds)

    try:
        wb.save(OUT_XLSX)
        print(f"Done!  Output: {OUT_XLSX}")
    except PermissionError:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_path = OUT_XLSX.parent / f"family_offices_v2_{timestamp}.xlsx"
        print(f"\n[WARN] Permission denied writing to {OUT_XLSX} (is it open in Excel?).")
        print(f"Saving to fallback file instead: {fallback_path}")
        wb.save(fallback_path)
        print(f"Done!  Output: {fallback_path}")
        
    print(f"  Records : {len(records)}")
    print(f"  Success : {len(records) - len(failed_crds)}")
    print(f"  Failed  : {len(failed_crds)}  {failed_crds}")


if __name__ == "__main__":
    main()
